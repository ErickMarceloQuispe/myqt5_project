import sqlite3
import sys
import os

from openpyxl import Workbook,load_workbook
from PyQt5.QtWidgets import QApplication, QWidget,QMainWindow
from PyQt5 import uic
from PyQt5 import QtGui,QtCore 
from PyQt5.QtGui import QCursor 

DB_FILE_NAME="raws_numbers.db"
TABLE_NAME="raws_numbers"
UI_FILE_NAME="ui_main.ui"

FOLDER_NAME="Archivos/"

EXTENSIONS=[".xlsx"]
LABEL_FILE_CB="--Seleccionar Archivo--"
LABEL_SHEET_CB="--Seleccionar Hoja--"
LABEL_TYPE_CB="--Seleccionar Tipo de Documento--"
LABEL_WAITING="En espera"

#Used to get raws of the files - Needs Manual Marks
def get_raws(input_sheet,end_row):
    raw_range=input_sheet["A1:A"+str(end_row)]
    raws_numbers=[]

    for i in range(len(raw_range)):
        if(not raw_range[i][0].value==(None)):
            raws_numbers.append(i+1)
    print(raws_numbers)
    return raws_numbers

#Get cell's values of a raw range
def get_data(raw_range):
    cols=get_cols_with_info(raw_range)
    data=[]
    for row in raw_range:
        aux=[]
        for col in cols:
            info=row[col].value
            #Agregar 0s en los espacios vacíos detectados
            #if(info==None):
            #    aux.append(0)
            #else:
            #    aux.append(info)
            aux.append(info)
        data.append(aux)
    return(data)

#
def print_data(data):
    for raw in data:
        text_raw=""
        for info in raw:
            text_raw=text_raw+"|"+str(info)
        print(text_raw)


def get_cols_with_info(raw_range):
    cols=[]
    for i in range(len(raw_range)):
        for j in range(len(raw_range[i])):
            if(not raw_range[i][j].value==(None)):
                if(j not in cols):
                    cols.append(j)
    cols.sort()
    return cols

def save_results(output_sheet,data,start_row):
    for i in range(len(data)):
        for j in range(len(data[i])):
            output_sheet.cell(row=start_row+i, column=j+1,value=data[i][j])
    return

def grupal_formate_date(arr,input_sheet,output_sheet):
    for i in range(0,len(arr),2):
        format_data_range(input_sheet,output_sheet,arr[i],arr[i+1])

def format_data_range(input_sheet,output_sheet,start_row,end_row):
    raw_range=input_sheet[str(start_row)+":"+str(end_row)]
    save_results(output_sheet,get_data(raw_range),start_row)
    return

class AppDemo(QMainWindow):
    def __init__(self):
        super().__init__()
        uic.loadUi(UI_FILE_NAME,self)

        self.input_book=None
        self.output_book=None
        self.input_sheet=None
        self.output_sheet=None
        self.raws_arr=None

        conn = sqlite3.connect(DB_FILE_NAME)
        c=conn.cursor()

        self.arrays=self.getArrsObject(c)
        conn.commit()
        conn.close()

        #SQLite
        self.comboBoxFilesIn.addItem(LABEL_FILE_CB)
        self.comboBoxFilesIn.addItems(self.get_file_names())
        self.comboBoxFilesIn.currentIndexChanged.connect(lambda:self.setFileIn())
        self.comboBoxSheetsIn.currentIndexChanged.connect(lambda:self.setSheet())

        self.comboBoxFilesOut.addItem(LABEL_FILE_CB)
        self.comboBoxFilesOut.addItems(self.get_file_names())
        self.comboBoxFilesOut.currentIndexChanged.connect(lambda:self.setFileOut())
        self.comboBoxSheetsOut.currentIndexChanged.connect(lambda:self.setSheet())
    
        self.comboBoxTypes.addItem(LABEL_TYPE_CB)
        for key in self.arrays.keys():
            self.comboBoxTypes.addItem(key)
        self.comboBoxTypes.currentIndexChanged.connect(lambda:self.setType())

        self.button.clicked.connect(lambda:self.startApp())

    def setFileIn(self):
        self.status_label.setText(LABEL_WAITING)
        if not self.input_sheet==None:
            self.input_sheet=None
        if(self.comboBoxFilesIn.currentText()==LABEL_FILE_CB):
            self.comboBoxSheetsIn.clear()
            return
        self.input_book=load_workbook(filename=FOLDER_NAME+self.comboBoxFilesIn.currentText())
        self.get_sheet_names(self.input_book,self.comboBoxSheetsIn)

    def setFileOut(self):
        self.status_label.setText(LABEL_WAITING)
        if not self.output_sheet==None:
            self.output_sheet=None
        if(self.comboBoxFilesOut.currentText()==LABEL_FILE_CB):
            self.comboBoxSheetsOut.clear()
            return
        self.output_book=load_workbook(filename=FOLDER_NAME+self.comboBoxFilesOut.currentText())
        self.get_sheet_names(self.output_book,self.comboBoxSheetsOut)
    
    def setSheet(self):
        self.status_label.setText(LABEL_WAITING)
        if (not self.input_book==None)and(not self.comboBoxSheetsIn==None):
            if(not self.comboBoxSheetsIn.currentText()=="")and(not self.comboBoxSheetsIn.currentText()==LABEL_SHEET_CB):
                self.input_sheet=self.input_book[self.comboBoxSheetsIn.currentText()]
        if (not self.output_book==None)and(not self.comboBoxSheetsOut==None):
            if(not self.comboBoxSheetsOut.currentText()=="")and(not self.comboBoxSheetsOut.currentText()==LABEL_SHEET_CB):
                self.output_sheet=self.output_book[self.comboBoxSheetsOut.currentText()]

    def startApp(self):
        isLessInfo=(self.input_book==None)or(self.output_book==None)or(self.input_sheet==None)or(self.output_sheet==None)or(self.raws_arr==None)
        if(isLessInfo):
            self.status_label.setText("Información Faltante")
            return    
        grupal_formate_date(self.raws_arr,self.input_sheet,self.output_sheet)
        self.output_book.save(FOLDER_NAME+self.comboBoxFilesOut.currentText())
        self.status_label.setText("Proceso Terminado")
        self.input_sheet=None
        self.output_sheet=None

    def setType(self):
        self.status_label.setText(LABEL_WAITING)
        if(self.comboBoxTypes.currentText()==LABEL_TYPE_CB):
            self.raws_arr=None
        else:
            self.raws_arr=self.arrays[self.comboBoxTypes.currentText()]

    def get_file_names(self):
        file_list=[]
        for root, directories, filenames in os.walk("./"+FOLDER_NAME):
            for filename in filenames:
                if(any(ext in filename for ext in EXTENSIONS)):
                    file_list.append(filename)
        return file_list

    def get_sheet_names(self,book,combobox):
        combobox.clear()
        combobox.addItem(LABEL_SHEET_CB)
        if book!=None:
            combobox.addItems(book.sheetnames)
    
    def getArrsObject(self,cursor):
        cursor.execute(""" 
            SELECT * FROM '{TABLE_NAME}' ORDER BY type
        """)
        obj={}
        for item in cursor.fetchall():
            obj[item[0]]=[int(numeric_string) for numeric_string in item[1].split(",")]
        return(obj)

# input_sheet=load_workbook(filename="Archivos/Abr-May Metales Pesados GERSA 2021.xlsx")["Hoja1"]
# get_raws(input_sheet,460)
# raise SystemExit

if __name__ == "__main__":
    app=QApplication(sys.argv)

    demo=AppDemo()
    demo.show()

    try:
        sys.exit(app.exec_())
    except SystemExit:
        print("Cerrando Ventana")

