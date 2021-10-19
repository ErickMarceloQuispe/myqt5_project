import sys
import os

from openpyxl import Workbook,load_workbook
from PyQt5.QtWidgets import QApplication, QWidget,QMainWindow
from PyQt5 import uic
from PyQt5 import QtGui,QtCore 
from PyQt5.QtGui import QCursor 

EXTENSIONS=[".xlsx"]
LABEL_FILE_CB="--Seleccionar Archivo--"
LABEL_SHEET_CB="--Seleccionar Hoja--"
LABEL_TYPE_CB="--Seleccionar Tipo de Documento--"
LABEL_WAITING="En espera"

#RAWS_FORMATO_NIÑOS=[13, 16, 21, 29, 33, 45, 50, 58, 63, 71, 76, 86, 90, 114, 119, 135, 140, 150, 155, 159, 165, 166, 173, 180, 187, 194, 200, 201, 206, 209, 214, 215, 223, 229, 233, 238, 245, 246, 252, 255, 260, 262, 267, 284, 287, 291, 296, 305, 310, 315, 319, 323, 329, 331, 338, 349, 355, 357, 362, 364, 369, 371, 376, 387, 392, 394, 400, 401, 407, 408, 415, 419, 425, 428]
#RAWS_NUTRICION=[11, 20, 24, 27, 32, 37, 40, 58, 61, 74, 78, 98, 101, 130, 134, 163, 167, 182, 185, 193, 197,201,205,206,209, 216, 222, 234, 238, 244, 247, 253]

def get_raws(input_sheet,end_row):
    raw_range=input_sheet["A1:A"+str(end_row)]
    raws_numbers=[]

    for i in range(len(raw_range)):
        if(not raw_range[i][0].value==(None)):
            raws_numbers.append(i+1)
    print(raws_numbers)
    return raws_numbers

def get_data(raw_range):
    cols=get_cols_with_info(raw_range)
    data=[]
    for row in raw_range:
        aux=[]
        for col in cols:
            info=row[col].value
            #Agregar 0s en los espacios vacíos detectados
            #if(info==None):
            #    aux.append(0)
            #else:
            #    aux.append(info)
            aux.append(info)
        data.append(aux)
    return(data)

def print_data(data):
    for raw in data:
        text_raw=""
        for info in raw:
            text_raw=text_raw+"|"+str(info)
        print(text_raw)


def get_cols_with_info(raw_range):
    cols=[]
    for i in range(len(raw_range)):
        for j in range(len(raw_range[i])):
            if(not raw_range[i][j].value==(None)):
                if(j not in cols):
                    cols.append(j)
    cols.sort()
    return cols

def save_results(output_sheet,data,start_row):
    for i in range(len(data)):
        for j in range(len(data[i])):
            output_sheet.cell(row=start_row+i, column=j+1,value=data[i][j])
    return

def grupal_formate_date(arr,input_sheet,output_sheet):
    for i in range(0,len(arr),2):
        format_data_range(input_sheet,output_sheet,arr[i],arr[i+1])

def format_data_range(input_sheet,output_sheet,start_row,end_row):
    raw_range=input_sheet[str(start_row)+":"+str(end_row)]
    save_results(output_sheet,get_data(raw_range),start_row)
    return

class AppDemo(QMainWindow):
    def __init__(self):
        super().__init__()
        uic.loadUi("demo1.ui",self)

        self.input_book=None
        self.output_book=None
        self.input_sheet=None
        self.output_sheet=None
        self.raws_arr=None

        self.arrays={
            "Niños":[13, 16, 21, 29, 33, 45, 50, 58, 63, 71, 76, 86, 90, 114, 119, 135, 140, 150, 155, 159, 165, 166, 173, 180, 187, 194, 200, 201, 206, 209, 214, 215, 223, 229, 233, 238, 245, 246, 252, 255, 260, 262, 267, 284, 287, 291, 296, 305, 310, 315, 319, 323, 329, 331, 338, 349, 355, 357, 362, 364, 369, 371, 376, 387, 392, 394, 400, 401, 407, 408, 415, 419, 425, 428],
            "Nutricion":[11, 20, 24, 27, 32, 37, 40, 58, 61, 74, 78, 98, 101, 130, 134, 163, 167, 182, 185, 193, 197,201,205,206,209, 216, 222, 234, 238, 244, 247, 253]
        }
        self.comboBoxFilesIn.addItem(LABEL_FILE_CB)
        self.comboBoxFilesIn.addItems(self.get_file_names())
        self.comboBoxFilesIn.currentIndexChanged.connect(lambda:self.setFileIn())
        self.comboBoxSheetsIn.currentIndexChanged.connect(lambda:self.setSheet())

        self.comboBoxFilesOut.addItem(LABEL_FILE_CB)
        self.comboBoxFilesOut.addItems(self.get_file_names())
        self.comboBoxFilesOut.currentIndexChanged.connect(lambda:self.setFileOut())
        self.comboBoxSheetsOut.currentIndexChanged.connect(lambda:self.setSheet())
    
        self.comboBoxTypes.addItem(LABEL_TYPE_CB)
        for key in self.arrays.keys():
            self.comboBoxTypes.addItem(key)
        self.comboBoxTypes.currentIndexChanged.connect(lambda:self.setType())

        self.button.clicked.connect(lambda:self.startApp())

    def setFileIn(self):
        self.status_label.setText(LABEL_WAITING)
        if not self.input_sheet==None:
            self.input_sheet=None
        if(self.comboBoxFilesIn.currentText()==LABEL_FILE_CB):
            self.comboBoxSheetsIn.clear()
            return
        self.input_book=load_workbook(filename=self.comboBoxFilesIn.currentText())
        self.get_sheet_names(self.input_book,self.comboBoxSheetsIn)

    def setFileOut(self):
        self.status_label.setText(LABEL_WAITING)
        if not self.output_sheet==None:
            self.output_sheet=None
        if(self.comboBoxFilesOut.currentText()==LABEL_FILE_CB):
            self.comboBoxSheetsOut.clear()
            return
        self.output_book=load_workbook(filename=self.comboBoxFilesOut.currentText())
        self.get_sheet_names(self.output_book,self.comboBoxSheetsOut)
    
    def setSheet(self):
        self.status_label.setText(LABEL_WAITING)
        if (not self.input_book==None)and(not self.comboBoxSheetsIn==None):
            if(not self.comboBoxSheetsIn.currentText()=="")and(not self.comboBoxSheetsIn.currentText()==LABEL_SHEET_CB):
                self.input_sheet=self.input_book[self.comboBoxSheetsIn.currentText()]
        if (not self.output_book==None)and(not self.comboBoxSheetsOut==None):
            if(not self.comboBoxSheetsOut.currentText()=="")and(not self.comboBoxSheetsOut.currentText()==LABEL_SHEET_CB):
                self.output_sheet=self.output_book[self.comboBoxSheetsOut.currentText()]

    def startApp(self):
        isLessInfo=(self.input_book==None)or(self.output_book==None)or(self.input_sheet==None)or(self.output_sheet==None)or(self.raws_arr==None)
        if(isLessInfo):
            self.status_label.setText("Información Faltante")
            return    
        grupal_formate_date(self.raws_arr,self.input_sheet,self.output_sheet)
        self.output_book.save(self.comboBoxFilesOut.currentText())
        self.status_label.setText("Proceso Terminado")
        self.input_sheet=None
        self.output_sheet=None

    def setType(self):
        self.status_label.setText(LABEL_WAITING)
        if(self.comboBoxTypes.currentText()==LABEL_TYPE_CB):
            self.raws_arr=None
        else:
            self.raws_arr=self.arrays[self.comboBoxTypes.currentText()]

    def get_file_names(self):
        file_list=[]
        for root, directories, filenames in os.walk("./"):
            for filename in filenames:
                if(any(ext in filename for ext in EXTENSIONS)):
                    file_list.append(filename)
        return file_list

    def get_sheet_names(self,book,combobox):
        combobox.clear()
        combobox.addItem(LABEL_SHEET_CB)
        if book!=None:
            combobox.addItems(book.sheetnames)

if __name__ == "__main__":
    app=QApplication(sys.argv)

    demo=AppDemo()
    demo.show()

    try:
        sys.exit(app.exec_())
    except SystemExit:
        print("Cerrando Ventana")