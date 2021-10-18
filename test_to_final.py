import sys
import os
from PyQt5.QtWidgets import * 
from PyQt5.QtGui import QPixmap 
from PyQt5 import QtGui,QtCore 
from PyQt5.QtGui import QCursor 

from openpyxl import Workbook,load_workbook


widgets={
    "combo_box_files":[],
    "combo_box_sheets":[],
    }

app = QApplication(sys.argv)
window=QWidget()
window.setWindowTitle("Wooooow")
window.setFixedWidth(1000)
#window.move(2700,200)
window.setStyleSheet("background: #161219;")

grid=QGridLayout()

def clear_widgets():
    for widget in widgets:
        if widgets[widget] != []:
            widgets[widget][-1].hide()
        for i in range(0,len(widgets[widget])):
            widgets[widget].pop()

def frame1():
    cb = QComboBox()
    cb.setStyleSheet(
        "font-size: 15px;"+
        "color:'white';")
    cb.addItems(get_file_names())
    cb.currentIndexChanged.connect(lambda:get_sheet_names(widgets["combo_box_files"][-1].currentText()))
    widgets["combo_box_files"].append(cb)
    grid.addWidget(cb)

EXTENSIONS=[".xlsx"]

def get_file_names():
    file_list=[]
    for root, directories, filenames in os.walk("./"):
        for filename in filenames:
            if(any(ext in filename for ext in EXTENSIONS)):
                file_list.append(filename)
    return file_list

def get_sheet_names(file_name):
    if(len(widgets["combo_box_sheets"])!=0):
        widgets["combo_box_sheets"].pop()
    input_book=load_workbook(filename=file_name)
    print()
    cb = QComboBox()
    cb.setStyleSheet(
        "font-size: 15px;"+
        "color:'white';")
    cb.addItems(input_book.sheetnames)
    cb.currentIndexChanged.connect(lambda:get_sheet_names(widgets["combo_box_sheets"][-1].currentText()))
    widgets["combo_box_sheets"].append(cb)
    grid.addWidget(cb)


get_file_names()
window.setLayout(grid)
frame1()
window.show()
sys.exit(app.exec())